from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill


matplotlib.use("Agg")

import seaborn as sns
from matplotlib import pyplot as plt
from matplotlib.lines import Line2D


ARQUIVO_ENTRADA = Path("CPC2021_CC_Bacharelado_Limpa.xlsx")
ARQUIVO_EXCEL_SAIDA = Path("Analise_Exploratoria_CC.xlsx")
ARQUIVO_GRAFICO_CATADM = Path("catadm_dist.png")
ARQUIVO_GRAFICO_MODALID = Path("modalid_dist.png")
ARQUIVO_BOXPLOT = Path("boxplot_notace_modalid.png")

COLUNA_NOTACE = "NOTACE"
COLUNA_NOTAFG = "NOTAFG"
COLUNA_CATADM = "CATADM"
COLUNA_MODALID = "MODALID"


def criar_tabela_frequencia(df: pd.DataFrame, coluna: str) -> pd.DataFrame:
    """Cria tabela de frequencia absoluta e relativa para uma coluna."""
    frequencia_absoluta = df[coluna].value_counts(dropna=False)
    frequencia_relativa = df[coluna].value_counts(normalize=True, dropna=False) * 100

    tabela = pd.DataFrame(
        {
            coluna: frequencia_absoluta.index,
            "Frequencia Absoluta": frequencia_absoluta.values,
            "Frequencia Relativa (%)": frequencia_relativa.values,
        }
    )

    return tabela


def plotar_barras(
    tabela: pd.DataFrame,
    coluna_categoria: str,
    arquivo_saida: Path,
    titulo: str,
    eixo_x: str,
    eixo_y: str,
) -> None:
    """Plota e salva grafico de barras com rotulos de valor."""
    plt.figure(figsize=(10, 6))

    barras = plt.bar(
        tabela[coluna_categoria].astype(str),
        tabela["Frequencia Absoluta"],
        color="#4F81BD",
        edgecolor="#1F4E78",
    )

    for barra in barras:
        altura = barra.get_height()
        plt.text(
            barra.get_x() + barra.get_width() / 2,
            altura,
            f"{int(altura)}",
            ha="center",
            va="bottom",
            fontsize=10,
        )

    plt.title(titulo, fontsize=14, fontweight="bold")
    plt.xlabel(eixo_x)
    plt.ylabel(eixo_y)
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()
    plt.savefig(arquivo_saida, dpi=300)
    plt.close()


def calcular_estatisticas(serie: pd.Series) -> pd.Series:
    """Calcula estatisticas descritivas para uma serie numerica."""
    serie = pd.to_numeric(serie, errors="coerce").dropna()
    q1 = serie.quantile(0.25)
    q3 = serie.quantile(0.75)
    media = serie.mean()
    desvio_padrao = serie.std(ddof=1)

    cv = np.nan
    if pd.notna(media) and media != 0:
        cv = (desvio_padrao / media) * 100

    return pd.Series(
        {
            "N": serie.count(),
            "Media": media,
            "Mediana": serie.median(),
            "Desvio Padrao": desvio_padrao,
            "Minimo": serie.min(),
            "Maximo": serie.max(),
            "Q1": q1,
            "Q3": q3,
            "IQR": q3 - q1,
            "CV(%)": cv,
        }
    )


def calcular_estatisticas_segmentadas(
    df: pd.DataFrame,
    coluna_valor: str,
    coluna_segmento: str,
) -> pd.DataFrame:
    """Calcula estatisticas descritivas por segmento."""
    estatisticas = (
        df.groupby(coluna_segmento, dropna=False)[coluna_valor]
        .apply(calcular_estatisticas)
        .unstack()
        .reset_index()
    )

    return estatisticas


def plotar_boxplot(df: pd.DataFrame) -> None:
    """Plota box-plot da NOTACE segmentado por modalidade."""
    plt.figure(figsize=(10, 6))

    sns.boxplot(
        data=df,
        x=COLUNA_MODALID,
        y=COLUNA_NOTACE,
        color="#D9EAF7",
        showmeans=True,
        meanprops={
            "marker": "D",
            "markerfacecolor": "#C00000",
            "markeredgecolor": "#7F0000",
            "markersize": 7,
        },
    )

    legenda_media = Line2D(
        [0],
        [0],
        marker="D",
        color="w",
        label="Media",
        markerfacecolor="#C00000",
        markeredgecolor="#7F0000",
        markersize=7,
    )

    plt.title("Box-plot da NOTACE por Modalidade de Ensino", fontsize=14, fontweight="bold")
    plt.xlabel("Modalidade de Ensino")
    plt.ylabel("NOTACE")
    plt.grid(axis="y", linestyle="--", alpha=0.6)
    plt.legend(handles=[legenda_media])
    plt.tight_layout()
    plt.savefig(ARQUIVO_BOXPLOT, dpi=300)
    plt.close()


def formatar_cabecalhos_excel(caminho_excel: Path) -> None:
    """Aplica formatacao de cabecalho nas abas do arquivo Excel."""
    from openpyxl import load_workbook

    workbook = load_workbook(caminho_excel)
    azul_escuro = "1F4E78"
    branco = "FFFFFF"

    header_fill = PatternFill(
        fill_type="solid",
        start_color=azul_escuro,
        end_color=azul_escuro,
    )
    header_font = Font(bold=True, color=branco)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    workbook.save(caminho_excel)


def exportar_tabelas_excel(
    tabela_catadm: pd.DataFrame,
    tabela_modalid: pd.DataFrame,
    estatisticas_gerais: pd.DataFrame,
    estatisticas_segmentadas: pd.DataFrame,
) -> None:
    """Exporta as tabelas da analise em abas separadas do Excel."""
    with pd.ExcelWriter(ARQUIVO_EXCEL_SAIDA, engine="openpyxl") as writer:
        tabela_catadm.to_excel(writer, sheet_name="Distribuicao_CATADM", index=False)
        tabela_modalid.to_excel(writer, sheet_name="Distribuicao_MODALID", index=False)
        estatisticas_gerais.to_excel(writer, sheet_name="Estatisticas_Gerais", index=False)
        estatisticas_segmentadas.to_excel(
            writer,
            sheet_name="Estatisticas_Segmentadas",
            index=False,
        )

    formatar_cabecalhos_excel(ARQUIVO_EXCEL_SAIDA)


def main() -> None:
    # PASSO 0: carrega a base e remove registros sem NOTACE ou NOTAFG.
    df = pd.read_excel(ARQUIVO_ENTRADA)
    total_inicial = len(df)
    df = df.dropna(subset=[COLUNA_NOTACE, COLUNA_NOTAFG]).reset_index(drop=True)
    total_final = len(df)
    removidas = total_inicial - total_final

    print("PASSO 0 - Limpeza previa")
    print(f"Linhas removidas: {removidas}")
    print(f"Linhas restantes: {total_final}")

    # PASSO 1: gera tabelas de frequencia e graficos de barras.
    tabela_catadm = criar_tabela_frequencia(df, COLUNA_CATADM)
    tabela_modalid = criar_tabela_frequencia(df, COLUNA_MODALID)

    plotar_barras(
        tabela=tabela_catadm,
        coluna_categoria=COLUNA_CATADM,
        arquivo_saida=ARQUIVO_GRAFICO_CATADM,
        titulo="Distribuicao por Categoria Administrativa",
        eixo_x="Categoria Administrativa",
        eixo_y="Frequencia Absoluta",
    )

    plotar_barras(
        tabela=tabela_modalid,
        coluna_categoria=COLUNA_MODALID,
        arquivo_saida=ARQUIVO_GRAFICO_MODALID,
        titulo="Distribuicao por Modalidade de Ensino",
        eixo_x="Modalidade de Ensino",
        eixo_y="Frequencia Absoluta",
    )

    # PASSO 2: calcula estatisticas descritivas gerais e por segmento.
    estatisticas_gerais = calcular_estatisticas(df[COLUNA_NOTACE]).to_frame().T
    estatisticas_gerais.insert(0, "Segmento", "Geral")

    coluna_segmento = COLUNA_MODALID
    if df[COLUNA_MODALID].nunique(dropna=False) == 1:
        coluna_segmento = COLUNA_CATADM

    estatisticas_segmentadas = calcular_estatisticas_segmentadas(
        df=df,
        coluna_valor=COLUNA_NOTACE,
        coluna_segmento=coluna_segmento,
    )
    estatisticas_segmentadas = estatisticas_segmentadas.rename(
        columns={coluna_segmento: "Segmento"}
    )
    estatisticas_segmentadas.insert(0, "Variavel de Segmentacao", coluna_segmento)

    print("\nPASSO 1 - Distribuicao por CATADM")
    print(tabela_catadm)

    print("\nPASSO 1 - Distribuicao por MODALID")
    print(tabela_modalid)

    print("\nPASSO 2 - Estatisticas descritivas gerais da NOTACE")
    print(estatisticas_gerais)

    print(f"\nPASSO 2 - Estatisticas da NOTACE por {coluna_segmento}")
    print(estatisticas_segmentadas)

    # PASSO 3: gera box-plot da NOTACE por modalidade.
    plotar_boxplot(df)

    # PASSO 4: exporta todas as tabelas em abas separadas no Excel.
    exportar_tabelas_excel(
        tabela_catadm=tabela_catadm,
        tabela_modalid=tabela_modalid,
        estatisticas_gerais=estatisticas_gerais,
        estatisticas_segmentadas=estatisticas_segmentadas,
    )

    print("\nArquivos gerados:")
    print(f"- {ARQUIVO_GRAFICO_CATADM}")
    print(f"- {ARQUIVO_GRAFICO_MODALID}")
    print(f"- {ARQUIVO_BOXPLOT}")
    print(f"- {ARQUIVO_EXCEL_SAIDA}")


if __name__ == "__main__":
    main()
