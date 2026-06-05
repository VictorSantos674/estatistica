from __future__ import annotations

import math
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = Path.home() / "Downloads" / "Dados_Dieta.xlsx"
OUTPUT_PATH = BASE_DIR / "outputs" / "dados_dieta_resolvido.xlsx"
PLOTS_DIR = BASE_DIR / "plots"


def read_data(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name="Sheet1")
    df = raw.iloc[:, :6].dropna(subset=[raw.columns[0]]).copy()
    df.columns = ["Id", "Idade", "Peso_Antes", "Peso_Apos", "Difpeso", "TipoAtivFis"]
    numeric_cols = ["Id", "Idade", "Peso_Antes", "Peso_Apos", "Difpeso"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def fit_ols(df: pd.DataFrame, y_col: str, x_cols: list[str]) -> dict[str, object]:
    y = df[y_col].to_numpy(dtype=float)
    x = df[x_cols].to_numpy(dtype=float)
    x = np.column_stack([np.ones(len(df)), x])
    coef = np.linalg.lstsq(x, y, rcond=None)[0]
    fitted = x @ coef
    residuals = y - fitted
    n = len(y)
    p = x.shape[1]
    df_resid = n - p
    sse = float(np.sum(residuals**2))
    sst = float(np.sum((y - np.mean(y)) ** 2))
    ssr = sst - sse
    mse = sse / df_resid
    r2 = 1.0 - sse / sst
    r2_adj = 1.0 - (1.0 - r2) * (n - 1) / df_resid
    rmse = math.sqrt(sse / n)
    erro_padrao_reg = math.sqrt(mse)
    f_stat = (ssr / (p - 1)) / mse

    xtx_inv = np.linalg.inv(x.T @ x)
    se = np.sqrt(np.diag(mse * xtx_inv))
    t_stat = coef / se

    return {
        "coef": coef,
        "se": se,
        "t": t_stat,
        "fitted": fitted,
        "residuals": residuals,
        "r2": r2,
        "r2_adj": r2_adj,
        "rmse": rmse,
        "erro_padrao_reg": erro_padrao_reg,
        "f_stat": f_stat,
        "df_resid": df_resid,
        "terms": ["Intercepto", *x_cols],
    }


def equation(model: dict[str, object], x_cols: list[str]) -> str:
    coef = model["coef"]
    parts = [f"Peso_Apos = {coef[0]:.6f}"]
    for name, value in zip(x_cols, coef[1:]):
        sign = "+" if value >= 0 else "-"
        parts.append(f"{sign} {abs(value):.6f}*{name}")
    return " ".join(parts)


def make_plots(df: pd.DataFrame) -> dict[str, Path]:
    PLOTS_DIR.mkdir(exist_ok=True)
    sns.set_theme(style="whitegrid", context="notebook")

    before_after = PLOTS_DIR / "dieta_peso_antes_depois.png"
    long_df = df.melt(
        id_vars=["Id"],
        value_vars=["Peso_Antes", "Peso_Apos"],
        var_name="Momento",
        value_name="Peso",
    )
    long_df["Momento"] = long_df["Momento"].replace(
        {"Peso_Antes": "Antes da dieta", "Peso_Apos": "Depois da dieta"}
    )
    fig, ax = plt.subplots(figsize=(7.5, 4.8), dpi=160)
    sns.boxplot(data=long_df, x="Momento", y="Peso", color="#dbeafe", ax=ax)
    sns.stripplot(data=long_df, x="Momento", y="Peso", color="#1d4ed8", size=5, alpha=0.75, ax=ax)
    ax.set_title("Peso antes e depois da dieta", weight="bold")
    ax.set_xlabel("")
    ax.set_ylabel("Peso (kg)")
    fig.tight_layout()
    fig.savefig(before_after, bbox_inches="tight")
    plt.close(fig)

    paired = PLOTS_DIR / "dieta_linhas_pareadas.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.8), dpi=160)
    for _, row in df.iterrows():
        ax.plot(["Antes", "Depois"], [row["Peso_Antes"], row["Peso_Apos"]], marker="o", alpha=0.55)
    ax.set_title("Mudanca individual de peso", weight="bold")
    ax.set_ylabel("Peso (kg)")
    fig.tight_layout()
    fig.savefig(paired, bbox_inches="tight")
    plt.close(fig)

    by_activity = PLOTS_DIR / "dieta_difpeso_por_atividade.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.8), dpi=160)
    order = ["Alta", "Media", "Baixa"]
    sns.barplot(
        data=df,
        x="TipoAtivFis",
        y="Difpeso",
        hue="TipoAtivFis",
        order=order,
        hue_order=order,
        errorbar=None,
        palette="Set2",
        legend=False,
        ax=ax,
    )
    sns.stripplot(data=df, x="TipoAtivFis", y="Difpeso", order=order, color="#111827", size=5, alpha=0.75, ax=ax)
    ax.axhline(0, color="#991b1b", linewidth=1.2)
    ax.set_title("Perda/ganho de peso por atividade fisica", weight="bold")
    ax.set_xlabel("Tipo de atividade fisica")
    ax.set_ylabel("Difpeso = peso antes - peso depois (kg)")
    fig.tight_layout()
    fig.savefig(by_activity, bbox_inches="tight")
    plt.close(fig)

    return {
        "before_after": before_after,
        "paired": paired,
        "by_activity": by_activity,
    }


def write_df(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def add_text(writer: pd.ExcelWriter, sheet_name: str, lines: list[str]) -> None:
    pd.DataFrame({"Analise": lines}).to_excel(writer, sheet_name=sheet_name, index=False)


def format_excel(path: Path, images: dict[str, list[tuple[Path, str]]]) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column_cells in ws.columns:
            width = min(max(max(len(str(c.value)) if c.value is not None else 0 for c in column_cells) + 2, 12), 55)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for sheet_name, sheet_images in images.items():
        ws = wb[sheet_name]
        for image_path, anchor in sheet_images:
            img = ExcelImage(str(image_path))
            img.width = 720
            img.height = 460
            ws.add_image(img, anchor)
    wb.save(path)


def solve() -> None:
    OUTPUT_PATH.parent.mkdir(exist_ok=True)
    df = read_data(INPUT_PATH)
    plots = make_plots(df)

    dist = df["TipoAtivFis"].value_counts().rename_axis("TipoAtivFis").reset_index(name="Frequencia")
    dist["Percentual"] = dist["Frequencia"] / len(df) * 100
    dist = dist.sort_values("TipoAtivFis").reset_index(drop=True)

    desc = df[["Peso_Antes", "Peso_Apos", "Difpeso"]].describe().T.reset_index()
    desc = desc.rename(columns={"index": "Variavel"})

    activity = (
        df.groupby("TipoAtivFis")
        .agg(
            Frequencia=("Difpeso", "count"),
            Media_Difpeso=("Difpeso", "mean"),
            Mediana_Difpeso=("Difpeso", "median"),
            Desvio_Padrao_Difpeso=("Difpeso", "std"),
            Min_Difpeso=("Difpeso", "min"),
            Max_Difpeso=("Difpeso", "max"),
            Media_Peso_Antes=("Peso_Antes", "mean"),
            Media_Peso_Apos=("Peso_Apos", "mean"),
        )
        .reset_index()
    )

    simple = fit_ols(df, "Peso_Apos", ["Peso_Antes"])
    multiple = fit_ols(df, "Peso_Apos", ["Idade", "Peso_Antes"])
    models = pd.DataFrame(
        [
            {
                "Modelo": "Simples: Peso_Apos ~ Peso_Antes",
                "Equacao": equation(simple, ["Peso_Antes"]),
                "R2": simple["r2"],
                "R2_ajustado": simple["r2_adj"],
                "RMSE": simple["rmse"],
                "Erro_padrao_regressao": simple["erro_padrao_reg"],
                "F": simple["f_stat"],
                "gl_residuo": simple["df_resid"],
                "Comentario": "Escolhido: maior R2 ajustado, menor erro padrao e modelo mais parcimonioso.",
            },
            {
                "Modelo": "Multiplo: Peso_Apos ~ Idade + Peso_Antes",
                "Equacao": equation(multiple, ["Idade", "Peso_Antes"]),
                "R2": multiple["r2"],
                "R2_ajustado": multiple["r2_adj"],
                "RMSE": multiple["rmse"],
                "Erro_padrao_regressao": multiple["erro_padrao_reg"],
                "F": multiple["f_stat"],
                "gl_residuo": multiple["df_resid"],
                "Comentario": "Nao escolhido: R2 bruto aumenta pouco, mas o R2 ajustado cai; idade acrescenta pouco.",
            },
        ]
    )

    coef_simple = pd.DataFrame(
        {
            "Termo": simple["terms"],
            "Coeficiente": simple["coef"],
            "Erro_padrao": simple["se"],
            "t": simple["t"],
        }
    )
    coef_multiple = pd.DataFrame(
        {
            "Termo": multiple["terms"],
            "Coeficiente": multiple["coef"],
            "Erro_padrao": multiple["se"],
            "t": multiple["t"],
        }
    )

    peso_antes = 75.7
    pred = float(simple["coef"][0] + simple["coef"][1] * peso_antes)
    prediction = pd.DataFrame(
        {
            "Modelo_escolhido": ["Simples: Peso_Apos ~ Peso_Antes"],
            "Peso_antes_informado": [peso_antes],
            "Formula": [equation(simple, ["Peso_Antes"])],
            "Previsao_Peso_Apos": [pred],
        }
    )

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        write_df(writer, "Dados", df)
        write_df(writer, "Q1 Distribuicao", dist)
        write_df(writer, "Descritivas", desc)
        add_text(
            writer,
            "Q2 Analise Dieta",
            [
                "A media do peso antes foi 59,69 kg e a media do peso depois foi 55,31 kg.",
                "A diferenca media foi 4,38 kg, indicando perda media de peso no grupo.",
                "A mediana caiu de 57,90 kg para 51,50 kg; graficamente a distribuicao depois fica mais baixa.",
                "Assim, a dieta aparenta ter sido eficaz para a amostra, embora nem todos tenham perdido peso.",
            ],
        )
        write_df(writer, "Q3 Perda por Ativ", activity)
        add_text(
            writer,
            "Q3 Comentario",
            [
                "Alta atividade fisica teve a maior perda media: 8,54 kg.",
                "Media atividade teve perda media positiva: 3,26 kg.",
                "Baixa atividade apresentou media negativa: -1,70 kg, ou seja, ganho medio de peso.",
                "A perda parece crescer com o impacto/intensidade da atividade fisica.",
            ],
        )
        write_df(writer, "Q4 Modelos", models)
        write_df(writer, "Coef Simples", coef_simple)
        write_df(writer, "Coef Multiplo", coef_multiple)
        add_text(
            writer,
            "Q4 Comentario",
            [
                "O modelo simples e preferivel.",
                "O modelo multiplo tem R2 ligeiramente maior, mas menor R2 ajustado e maior erro padrao.",
                "A variavel Idade nao melhora suficientemente o ajuste; portanto o modelo simples e mais parcimonioso.",
            ],
        )
        write_df(writer, "Q5 Previsao", prediction)
        add_text(
            writer,
            "Graficos",
            [
                "Graficos incluidos nesta aba: pesos antes/depois, linhas pareadas individuais e Difpeso por atividade fisica.",
            ],
        )

    format_excel(
        OUTPUT_PATH,
        {
            "Graficos": [
                (plots["before_after"], "A4"),
                (plots["paired"], "A28"),
                (plots["by_activity"], "A52"),
            ]
        },
    )

    print(f"Planilha gerada: {OUTPUT_PATH}")
    print(f"Grafico antes/depois: {plots['before_after']}")
    print(f"Grafico linhas pareadas: {plots['paired']}")
    print(f"Grafico por atividade: {plots['by_activity']}")
    print("\nDistribuicao:")
    print(dist.to_string(index=False))
    print("\nPerda por atividade:")
    print(activity.to_string(index=False))
    print("\nComparacao dos modelos:")
    print(models[["Modelo", "R2", "R2_ajustado", "Erro_padrao_regressao", "F"]].to_string(index=False))
    print(f"\nPrevisao para peso antes = {peso_antes:.1f} kg: {pred:.4f} kg")


if __name__ == "__main__":
    solve()
