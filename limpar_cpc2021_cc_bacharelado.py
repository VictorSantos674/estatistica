from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill


ARQUIVO_ORIGINAL = Path(r"C:\Users\3vict\Downloads\CPC_2021.xlsx")
ARQUIVO_SAIDA = Path("CPC2021_CC_Bacharelado_Limpa.xlsx")


COLUNAS_RENOMEADAS = {
    "Ano": "Ano",
    "Área de Avaliação": "Curso",
    "Grau acadêmico": "Grau Acadêmico",
    "Código da IES*": "Código IES",
    "Categoria Administrativa*": "CATADM",
    "Modalidade de Ensino***": "MODALID",
    " Nota Bruta - FG": "NOTAFG",
    " Nota Bruta - CE": "NOTACE",
    " Nota Bruta - Doutores": "DOUTORES",
}

COLUNAS_NUMERICAS = ["NOTAFG", "NOTACE", "DOUTORES"]


def validar_colunas(df: pd.DataFrame) -> None:
    colunas_faltantes = [
        coluna for coluna in COLUNAS_RENOMEADAS if coluna not in df.columns
    ]

    if colunas_faltantes:
        faltantes = "\n".join(f"- {coluna}" for coluna in colunas_faltantes)
        raise KeyError(
            "As seguintes colunas esperadas nao foram encontradas no arquivo:\n"
            f"{faltantes}"
        )


def formatar_planilha(caminho_saida: Path, df: pd.DataFrame) -> None:
    azul_escuro = "1F4E78"
    azul_claro = "D9EAF7"
    branco = "FFFFFF"

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CPC2021_CC")

        worksheet = writer.sheets["CPC2021_CC"]
        worksheet.freeze_panes = "A2"

        header_fill = PatternFill(
            fill_type="solid",
            start_color=azul_escuro,
            end_color=azul_escuro,
        )
        header_font = Font(bold=True, color=branco)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        white_fill = PatternFill(
            fill_type="solid",
            start_color=branco,
            end_color=branco,
        )
        light_blue_fill = PatternFill(
            fill_type="solid",
            start_color=azul_claro,
            end_color=azul_claro,
        )

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
        ):
            fill = white_fill if row[0].row % 2 == 0 else light_blue_fill
            for cell in row:
                cell.fill = fill

        for coluna in COLUNAS_NUMERICAS:
            col_idx = df.columns.get_loc(coluna) + 1
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                row[0].number_format = "0.0000"


def limpar_base() -> pd.DataFrame:
    df = pd.read_excel(ARQUIVO_ORIGINAL)

    validar_colunas(df)

    filtro_curso = df["Área de Avaliação"].str.contains(
        "CIÊNCIA DA COMPUTAÇÃO",
        case=False,
        na=False,
    )
    filtro_grau = df["Grau acadêmico"].str.contains(
        "BACHARELADO",
        case=False,
        na=False,
    )

    df_limpo = df.loc[filtro_curso & filtro_grau, list(COLUNAS_RENOMEADAS)]
    df_limpo = df_limpo.rename(columns=COLUNAS_RENOMEADAS)
    df_limpo = df_limpo.replace("-", np.nan)

    for coluna in COLUNAS_NUMERICAS:
        df_limpo[coluna] = pd.to_numeric(df_limpo[coluna], errors="coerce")

    df_limpo = df_limpo.reset_index(drop=True)

    return df_limpo


def main() -> None:
    df_limpo = limpar_base()
    formatar_planilha(ARQUIVO_SAIDA, df_limpo)

    print("Shape do DataFrame limpo:")
    print(df_limpo.shape)

    print("\nContagem de valores nulos por coluna:")
    print(df_limpo.isna().sum())

    print("\nPrimeiras 5 linhas:")
    print(df_limpo.head())


if __name__ == "__main__":
    main()
