from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = Path.home() / "Downloads" / "aula_regressaoC3 (1).xlsx"
DEFAULT_OUTPUT = BASE_DIR / "outputs" / "aula_regressaoC3_resolvida.xlsx"
PLOTS_DIR = BASE_DIR / "plots"


try:
    from scipy import stats as scipy_stats
except Exception:  # pragma: no cover - fallback for environments without scipy
    scipy_stats = None


@dataclass
class OLSResult:
    name: str
    predictors: list[str]
    coef: np.ndarray
    se: np.ndarray
    t_stat: np.ndarray
    p_values: np.ndarray
    fitted: np.ndarray
    residuals: np.ndarray
    r2: float
    r2_adj: float
    rmse: float
    f_stat: float
    f_pvalue: float
    mse: float
    df_resid: int
    x_columns: list[str]
    xtx_inv: np.ndarray


def _betacf(a: float, b: float, x: float) -> float:
    max_iter = 200
    eps = 3e-14
    fpmin = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < fpmin:
        d = fpmin
    d = 1.0 / d
    h = d

    for m in range(1, max_iter + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        h *= d * c

        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            break
    return h


def regularized_beta(a: float, b: float, x: float) -> float:
    if x <= 0.0:
        return 0.0
    if x >= 1.0:
        return 1.0
    bt = math.exp(
        math.lgamma(a + b)
        - math.lgamma(a)
        - math.lgamma(b)
        + a * math.log(x)
        + b * math.log(1.0 - x)
    )
    if x < (a + 1.0) / (a + b + 2.0):
        return bt * _betacf(a, b, x) / a
    return 1.0 - bt * _betacf(b, a, 1.0 - x) / b


def t_cdf(t_value: float, df: int) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.t.cdf(t_value, df))
    x = df / (df + t_value * t_value)
    ib = regularized_beta(df / 2.0, 0.5, x)
    if t_value >= 0:
        return 1.0 - 0.5 * ib
    return 0.5 * ib


def t_two_sided_pvalue(t_value: float, df: int) -> float:
    if scipy_stats is not None:
        return float(2.0 * scipy_stats.t.sf(abs(t_value), df))
    return float(2.0 * min(t_cdf(t_value, df), 1.0 - t_cdf(t_value, df)))


def t_ppf(probability: float, df: int) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.t.ppf(probability, df))

    low, high = -1.0, 1.0
    while t_cdf(low, df) > probability:
        low *= 2.0
    while t_cdf(high, df) < probability:
        high *= 2.0

    for _ in range(100):
        mid = (low + high) / 2.0
        if t_cdf(mid, df) < probability:
            low = mid
        else:
            high = mid
    return (low + high) / 2.0


def f_sf(f_value: float, df_model: int, df_resid: int) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.f.sf(f_value, df_model, df_resid))
    x = df_resid / (df_resid + df_model * f_value)
    return float(regularized_beta(df_resid / 2.0, df_model / 2.0, x))


def read_student_data(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=0)
    columns = ["Id", "Idade", "Horas", "Nota"]
    df = raw[columns].copy()
    for column in columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")
    return df.dropna(subset=columns).reset_index(drop=True)


def make_scatter_plot(df: pd.DataFrame, x: str, y: str, path: Path) -> None:
    sns.set_theme(style="whitegrid", context="notebook")
    fig, ax = plt.subplots(figsize=(7, 4.6), dpi=160)
    sns.regplot(
        data=df,
        x=x,
        y=y,
        ci=95,
        scatter_kws={"s": 50, "alpha": 0.85, "color": "#2563eb"},
        line_kws={"color": "#dc2626", "linewidth": 2},
        ax=ax,
    )
    ax.set_title(f"Dispersao: {y} x {x}", fontsize=13, weight="bold")
    ax.set_xlabel(x)
    ax.set_ylabel(y)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def fit_ols(df: pd.DataFrame, y_col: str, x_cols: list[str], name: str) -> OLSResult:
    y = df[y_col].to_numpy(dtype=float)
    x_raw = df[x_cols].to_numpy(dtype=float)
    x = np.column_stack([np.ones(len(df)), x_raw])

    coef = np.linalg.lstsq(x, y, rcond=None)[0]
    fitted = x @ coef
    residuals = y - fitted

    n = len(y)
    p = x.shape[1]
    df_model = p - 1
    df_resid = n - p
    sse = float(np.sum(residuals**2))
    sst = float(np.sum((y - np.mean(y)) ** 2))
    ssr = sst - sse
    mse = sse / df_resid
    rmse = math.sqrt(sse / n)
    r2 = 1.0 - sse / sst
    r2_adj = 1.0 - (1.0 - r2) * (n - 1) / df_resid
    f_stat = (ssr / df_model) / mse
    f_pvalue = f_sf(f_stat, df_model, df_resid)

    xtx_inv = np.linalg.inv(x.T @ x)
    covariance = mse * xtx_inv
    se = np.sqrt(np.diag(covariance))
    t_stat = coef / se
    p_values = np.array([t_two_sided_pvalue(value, df_resid) for value in t_stat])

    return OLSResult(
        name=name,
        predictors=x_cols,
        coef=coef,
        se=se,
        t_stat=t_stat,
        p_values=p_values,
        fitted=fitted,
        residuals=residuals,
        r2=r2,
        r2_adj=r2_adj,
        rmse=rmse,
        f_stat=f_stat,
        f_pvalue=f_pvalue,
        mse=mse,
        df_resid=df_resid,
        x_columns=["Intercepto", *x_cols],
        xtx_inv=xtx_inv,
    )


def coefficient_table(result: OLSResult) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Termo": result.x_columns,
            "Coeficiente": result.coef,
            "Erro padrao": result.se,
            "t": result.t_stat,
            "p-valor": result.p_values,
        }
    )


def metrics_table(result: OLSResult) -> pd.DataFrame:
    equation = equation_text(result)
    return pd.DataFrame(
        {
            "Metrica": [
                "Equacao ajustada",
                "R2",
                "R2 ajustado",
                "RMSE",
                "F-estatistica",
                "p-valor do F",
                "gl residuais",
            ],
            "Valor": [
                equation,
                result.r2,
                result.r2_adj,
                result.rmse,
                result.f_stat,
                result.f_pvalue,
                result.df_resid,
            ],
        }
    )


def equation_text(result: OLSResult) -> str:
    pieces = [f"Nota = {result.coef[0]:.6f}"]
    for predictor, coef in zip(result.predictors, result.coef[1:]):
        sign = "+" if coef >= 0 else "-"
        pieces.append(f"{sign} {abs(coef):.6f}*{predictor}")
    return " ".join(pieces)


def predict_with_interval(
    result: OLSResult, values: dict[str, float], confidence: float = 0.95
) -> pd.DataFrame:
    x0 = np.array([1.0, *[values[predictor] for predictor in result.predictors]])
    y_hat = float(x0 @ result.coef)
    leverage = float(x0 @ result.xtx_inv @ x0.T)
    critical = t_ppf(1.0 - (1.0 - confidence) / 2.0, result.df_resid)
    se_mean = math.sqrt(result.mse * leverage)
    se_prediction = math.sqrt(result.mse * (1.0 + leverage))

    return pd.DataFrame(
        {
            "Modelo": [result.name],
            "Valores usados": [", ".join(f"{key}={value}" for key, value in values.items())],
            "Predicao pontual": [y_hat],
            "IC95 media - inferior": [y_hat - critical * se_mean],
            "IC95 media - superior": [y_hat + critical * se_mean],
            "IP95 aluno - inferior": [y_hat - critical * se_prediction],
            "IP95 aluno - superior": [y_hat + critical * se_prediction],
        }
    )


def residual_table(df: pd.DataFrame, result: OLSResult) -> pd.DataFrame:
    table = df[["Id", "Idade", "Horas", "Nota"]].copy()
    table[f"Nota ajustada - {result.name}"] = result.fitted
    table[f"Residuo - {result.name}"] = result.residuals
    return table


def compare_models(results: list[OLSResult]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Modelo": [result.name for result in results],
            "Equacao": [equation_text(result) for result in results],
            "R2": [result.r2 for result in results],
            "R2 ajustado": [result.r2_adj for result in results],
            "RMSE": [result.rmse for result in results],
            "F-estatistica": [result.f_stat for result in results],
            "p-valor do F": [result.f_pvalue for result in results],
        }
    ).sort_values(["R2 ajustado", "RMSE"], ascending=[False, True])


def write_df(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    startrow: int = 0,
    startcol: int = 0,
    index: bool = False,
) -> None:
    df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=startrow,
        startcol=startcol,
        index=index,
    )


def add_text_sheet(writer: pd.ExcelWriter, sheet_name: str, lines: list[str]) -> None:
    pd.DataFrame({"Resposta": lines}).to_excel(writer, sheet_name=sheet_name, index=False)


def format_workbook(output_path: Path, image_map: dict[str, tuple[Path, str]]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            width = min(max(max_len + 2, 12), 48)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for sheet_name, (image_path, anchor) in image_map.items():
        ws = wb[sheet_name]
        img = ExcelImage(str(image_path))
        img.width = 720
        img.height = 470
        ws.add_image(img, anchor)

    wb.save(output_path)


def solve(input_path: Path, output_path: Path) -> None:
    PLOTS_DIR.mkdir(exist_ok=True)
    output_path.parent.mkdir(exist_ok=True)

    df = read_student_data(input_path)

    plot_idade = PLOTS_DIR / "regressao_nota_idade.png"
    plot_horas = PLOTS_DIR / "regressao_nota_horas.png"
    make_scatter_plot(df, "Idade", "Nota", plot_idade)
    make_scatter_plot(df, "Horas", "Nota", plot_horas)

    corr = df[["Nota", "Idade", "Horas"]].corr(method="pearson")

    model_horas = fit_ols(df, "Nota", ["Horas"], "Nota ~ Horas")
    model_idade = fit_ols(df, "Nota", ["Idade"], "Nota ~ Idade")
    model_multiplo = fit_ols(df, "Nota", ["Idade", "Horas"], "Nota ~ Idade + Horas")
    results = [model_horas, model_idade, model_multiplo]
    comparison = compare_models(results)
    best_simple = max([model_horas, model_idade], key=lambda item: item.r2_adj)
    best_general = comparison.iloc[0]["Modelo"]
    prediction_9h = predict_with_interval(model_horas, {"Horas": 9.0})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_df(writer, "Dados", df)

        add_text_sheet(
            writer,
            "Q1 Grafico Idade",
            [
                "Grafico de dispersao entre Nota e Idade com reta de tendencia.",
                "A inclinacao e baixa quando comparada ao grafico de Horas.",
            ],
        )
        add_text_sheet(
            writer,
            "Q2 Grafico Horas",
            [
                "Grafico de dispersao entre Nota e Horas com reta de tendencia.",
                "A relacao visual e positiva: mais horas de estudo tendem a acompanhar notas maiores.",
            ],
        )
        add_text_sheet(
            writer,
            "Q3 Analise Grafica",
            [
                f"Pela analise grafica, a variavel que melhor representa o desempenho e: {best_simple.predictors[0]}.",
                "O grafico Nota x Horas apresenta tendencia linear positiva mais evidente.",
            ],
        )

        corr.to_excel(writer, sheet_name="Q4 Correlacao")
        add_text_sheet(
            writer,
            "Q4 Resposta",
            [
                f"A maior correlacao absoluta com Nota confirma a escolha: {best_simple.predictors[0]}.",
                f"Correlacao Nota-Horas = {corr.loc['Nota', 'Horas']:.6f}.",
                f"Correlacao Nota-Idade = {corr.loc['Nota', 'Idade']:.6f}.",
            ],
        )

        write_df(writer, "Q5 Reg Horas", metrics_table(model_horas))
        write_df(writer, "Q5 Coef Horas", coefficient_table(model_horas))
        write_df(writer, "Q5 Residuos Horas", residual_table(df, model_horas))

        write_df(writer, "Q6 Reg Idade", metrics_table(model_idade))
        write_df(writer, "Q6 Coef Idade", coefficient_table(model_idade))
        write_df(writer, "Q6 Residuos Idade", residual_table(df, model_idade))

        write_df(writer, "Q7 Predicao 9h", prediction_9h)
        write_df(writer, "Q8 Reg Multipla", metrics_table(model_multiplo))
        write_df(writer, "Q8 Coef Multipla", coefficient_table(model_multiplo))
        write_df(writer, "Q8 Residuos Mult", residual_table(df, model_multiplo))
        write_df(writer, "Q9 Comparacao", comparison)
        add_text_sheet(
            writer,
            "Conclusao",
            [
                f"Entre os modelos simples, o melhor e {best_simple.name}, pois tem maior R2 ajustado e menor RMSE.",
                f"Na comparacao geral, o melhor pela tabela e {best_general}.",
                "Para explicar desempenho, Horas de estudo e a variavel mais relevante neste conjunto de dados.",
            ],
        )

    format_workbook(
        output_path,
        {
            "Q1 Grafico Idade": (plot_idade, "A5"),
            "Q2 Grafico Horas": (plot_horas, "A5"),
        },
    )

    print(f"Planilha gerada: {output_path}")
    print(f"Grafico Nota x Idade: {plot_idade}")
    print(f"Grafico Nota x Horas: {plot_horas}")
    print("\nResumo dos modelos:")
    print(comparison.to_string(index=False))
    print("\nPredicao para 9 horas:")
    print(prediction_9h.to_string(index=False))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Resolve exercicio de regressao linear.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Caminho da planilha original.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Caminho da planilha resolvida.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    solve(args.input, args.output)
