from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats


# ===========================================================
# CONFIGURACOES GERAIS DO PROJETO
# ===========================================================

ARQUIVO_ENTRADA = Path("CPC2021_CC_Bacharelado_Limpa.xlsx")
ARQUIVO_EXCEL_SAIDA = Path("CPC2021_CC_Regressao.xlsx")

ARQUIVO_HEATMAP = Path("correlacao_heatmap.png")
ARQUIVO_SCATTER_FG = Path("scatter_FG.png")
ARQUIVO_SCATTER_DOUTORES = Path("scatter_DOUTORES.png")
ARQUIVO_DIAGNOSTICO = Path("diagnostico_residuos.png")

COLUNAS_MODELO = ["NOTACE", "NOTAFG", "DOUTORES"]
DATA_ENTREGA = "29/05/2026"

AZUL_ESCURO = "1F4E79"
AZUL_CLARO = "BDD7EE"
BRANCO = "FFFFFF"
VERDE = "C6EFCE"
AMARELO = "FFF2CC"


@dataclass
class ResultadoRegressao:
    """Armazena todos os resultados necessarios para comparar e exportar modelos."""

    nome: str
    variaveis: list[str]
    equacao: str
    coeficientes: np.ndarray
    erros_padrao: np.ndarray
    t_estatisticas: np.ndarray
    p_valores: np.ndarray
    r2: float
    r2_ajustado: float
    rmse: float
    f_estatistica: float
    f_pvalor: float
    y_previsto: np.ndarray
    residuos: np.ndarray
    mse: float
    gl_residuais: int
    xtx_inv: np.ndarray
    nomes_coeficientes: list[str]


# ===========================================================
# PASSO 1 - CARREGAMENTO E LIMPEZA PREVIA
# ===========================================================


def carregar_e_limpar_dados() -> pd.DataFrame:
    """Carrega a base, remove valores ausentes em cascata e informa as perdas."""
    df = pd.read_excel(ARQUIVO_ENTRADA)

    shape_inicial = df.shape
    print(f"Shape inicial: {shape_inicial}")

    df_sem_notas = df.dropna(subset=["NOTACE", "NOTAFG"]).copy()
    removidas_notas = len(df) - len(df_sem_notas)
    print(f"Após remover NaN em NOTACE/NOTAFG: {df_sem_notas.shape}")
    print(f"Linhas removidas nessa etapa: {removidas_notas}")

    df_limpo = df_sem_notas.dropna(subset=["DOUTORES"]).copy()
    removidas_doutores = len(df_sem_notas) - len(df_limpo)
    print(f"Após remover NaN em DOUTORES: {df_limpo.shape}")
    print(f"Linhas removidas nessa etapa: {removidas_doutores}")
    print(f"Total de linhas removidas: {len(df) - len(df_limpo)}")

    for coluna in COLUNAS_MODELO:
        df_limpo[coluna] = pd.to_numeric(df_limpo[coluna], errors="coerce")

    df_limpo = df_limpo.dropna(subset=COLUNAS_MODELO).reset_index(drop=True)
    df_limpo.insert(0, "Id", np.arange(1, len(df_limpo) + 1))
    return df_limpo


# ===========================================================
# PASSO 2 - ANALISE DE CORRELACAO E GRAFICOS
# ===========================================================


def gerar_graficos_correlacao(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula a matriz de correlacao e salva heatmap e dispersoes."""
    matriz_correlacao = df[COLUNAS_MODELO].corr(method="pearson")

    sns.set_theme(style="whitegrid", font="Arial")

    plt.figure(figsize=(7, 5), dpi=150)
    sns.heatmap(
        matriz_correlacao,
        annot=True,
        fmt=".4f",
        cmap="Blues",
        linewidths=0.5,
        cbar=True,
    )
    plt.title("Matriz de Correlação de Pearson", fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(ARQUIVO_HEATMAP, dpi=300)
    plt.close()

    plotar_dispersao_com_tendencia(
        df,
        x_col="NOTAFG",
        y_col="NOTACE",
        arquivo_saida=ARQUIVO_SCATTER_FG,
        titulo="NOTACE x NOTAFG",
    )
    plotar_dispersao_com_tendencia(
        df,
        x_col="DOUTORES",
        y_col="NOTACE",
        arquivo_saida=ARQUIVO_SCATTER_DOUTORES,
        titulo="NOTACE x DOUTORES",
    )

    return matriz_correlacao


def plotar_dispersao_com_tendencia(
    df: pd.DataFrame,
    x_col: str,
    y_col: str,
    arquivo_saida: Path,
    titulo: str,
) -> None:
    """Plota dispersao e linha de tendencia linear calculada com numpy.polyfit."""
    x = df[x_col].to_numpy(dtype=float)
    y = df[y_col].to_numpy(dtype=float)
    coef_angular, intercepto = np.polyfit(x, y, deg=1)

    x_linha = np.linspace(x.min(), x.max(), 100)
    y_linha = coef_angular * x_linha + intercepto

    plt.figure(figsize=(7, 5), dpi=150)
    plt.scatter(x, y, color="#2F75B5", alpha=0.75, edgecolor="#1F4E79")
    plt.plot(x_linha, y_linha, color="#C00000", linewidth=2)
    plt.title(titulo, fontsize=13, fontweight="bold")
    plt.xlabel(x_col)
    plt.ylabel(y_col)
    plt.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(arquivo_saida, dpi=300)
    plt.close()


# ===========================================================
# PASSO 3 - AJUSTE DOS MODELOS DE REGRESSAO
# ===========================================================


def ajustar_modelo_simples(df: pd.DataFrame, x_col: str, nome: str) -> ResultadoRegressao:
    """Ajusta regressao simples com scipy.stats.linregress."""
    y = df["NOTACE"].to_numpy(dtype=float)
    x = df[x_col].to_numpy(dtype=float)
    n = len(y)

    resultado = stats.linregress(x, y)
    coeficientes = np.array([resultado.intercept, resultado.slope], dtype=float)

    x_design = np.column_stack([np.ones(n), x])
    y_previsto = x_design @ coeficientes
    residuos = y - y_previsto

    metricas = calcular_metricas_ols(y, y_previsto, residuos, x_design)
    erros_padrao, t_estatisticas, p_valores = calcular_inferencia_coeficientes(
        coeficientes,
        metricas["mse"],
        metricas["gl_residuais"],
        metricas["xtx_inv"],
    )

    equacao = (
        f"NOTACE = {coeficientes[0]:.4f} + "
        f"{coeficientes[1]:.4f}*{x_col}"
    )

    return ResultadoRegressao(
        nome=nome,
        variaveis=[x_col],
        equacao=equacao,
        coeficientes=coeficientes,
        erros_padrao=erros_padrao,
        t_estatisticas=t_estatisticas,
        p_valores=p_valores,
        r2=metricas["r2"],
        r2_ajustado=metricas["r2_ajustado"],
        rmse=metricas["rmse"],
        f_estatistica=metricas["f_estatistica"],
        f_pvalor=metricas["f_pvalor"],
        y_previsto=y_previsto,
        residuos=residuos,
        mse=metricas["mse"],
        gl_residuais=metricas["gl_residuais"],
        xtx_inv=metricas["xtx_inv"],
        nomes_coeficientes=["Intercepto", x_col],
    )


def ajustar_modelo_multiplo(df: pd.DataFrame) -> ResultadoRegressao:
    """Ajusta regressao multipla usando numpy.linalg.lstsq."""
    y = df["NOTACE"].to_numpy(dtype=float)
    x_design = np.column_stack(
        [
            np.ones(len(df)),
            df["NOTAFG"].to_numpy(dtype=float),
            df["DOUTORES"].to_numpy(dtype=float),
        ]
    )

    coeficientes = np.linalg.lstsq(x_design, y, rcond=None)[0]
    y_previsto = x_design @ coeficientes
    residuos = y - y_previsto

    metricas = calcular_metricas_ols(y, y_previsto, residuos, x_design)
    erros_padrao, t_estatisticas, p_valores = calcular_inferencia_coeficientes(
        coeficientes,
        metricas["mse"],
        metricas["gl_residuais"],
        metricas["xtx_inv"],
    )

    equacao = (
        f"NOTACE = {coeficientes[0]:.4f} + "
        f"{coeficientes[1]:.4f}*NOTAFG + "
        f"{coeficientes[2]:.4f}*DOUTORES"
    )

    return ResultadoRegressao(
        nome="Modelo 3 - Multipla",
        variaveis=["NOTAFG", "DOUTORES"],
        equacao=equacao,
        coeficientes=coeficientes,
        erros_padrao=erros_padrao,
        t_estatisticas=t_estatisticas,
        p_valores=p_valores,
        r2=metricas["r2"],
        r2_ajustado=metricas["r2_ajustado"],
        rmse=metricas["rmse"],
        f_estatistica=metricas["f_estatistica"],
        f_pvalor=metricas["f_pvalor"],
        y_previsto=y_previsto,
        residuos=residuos,
        mse=metricas["mse"],
        gl_residuais=metricas["gl_residuais"],
        xtx_inv=metricas["xtx_inv"],
        nomes_coeficientes=["Intercepto", "NOTAFG", "DOUTORES"],
    )


def calcular_metricas_ols(
    y: np.ndarray,
    y_previsto: np.ndarray,
    residuos: np.ndarray,
    x_design: np.ndarray,
) -> dict[str, float | int | np.ndarray]:
    """Calcula R2, R2 ajustado, RMSE, F e matriz (X'X)^-1."""
    n = len(y)
    p = x_design.shape[1]
    k = p - 1
    gl_residuais = n - p

    sse = float(np.sum(residuos**2))
    sst = float(np.sum((y - np.mean(y)) ** 2))
    ssr = sst - sse

    mse = sse / gl_residuais
    rmse = float(np.sqrt(sse / n))
    r2 = 1.0 - (sse / sst)
    r2_ajustado = 1.0 - (1.0 - r2) * (n - 1) / gl_residuais
    f_estatistica = (ssr / k) / mse
    f_pvalor = float(stats.f.sf(f_estatistica, k, gl_residuais))
    xtx_inv = np.linalg.inv(x_design.T @ x_design)

    return {
        "r2": float(r2),
        "r2_ajustado": float(r2_ajustado),
        "rmse": rmse,
        "f_estatistica": float(f_estatistica),
        "f_pvalor": f_pvalor,
        "mse": float(mse),
        "gl_residuais": gl_residuais,
        "xtx_inv": xtx_inv,
    }


def calcular_inferencia_coeficientes(
    coeficientes: np.ndarray,
    mse: float,
    gl_residuais: int,
    xtx_inv: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Calcula erros padrao, estatisticas t e p-valores dos coeficientes."""
    matriz_covariancia = mse * xtx_inv
    erros_padrao = np.sqrt(np.diag(matriz_covariancia))
    t_estatisticas = coeficientes / erros_padrao
    p_valores = 2.0 * stats.t.sf(np.abs(t_estatisticas), gl_residuais)
    return erros_padrao, t_estatisticas, p_valores


def ajustar_todos_os_modelos(df: pd.DataFrame) -> list[ResultadoRegressao]:
    """Executa os tres modelos pedidos no enunciado."""
    return [
        ajustar_modelo_simples(df, "NOTAFG", "Modelo 1 - Simples NOTAFG"),
        ajustar_modelo_simples(df, "DOUTORES", "Modelo 2 - Simples DOUTORES"),
        ajustar_modelo_multiplo(df),
    ]


# ===========================================================
# PASSO 4 - TABELA COMPARATIVA DOS MODELOS
# ===========================================================


def criar_tabela_comparacao(modelos: list[ResultadoRegressao]) -> pd.DataFrame:
    """Cria a tabela comparativa com as principais metricas dos modelos."""
    return pd.DataFrame(
        {
            "Modelo": [modelo.nome for modelo in modelos],
            "Variaveis": [", ".join(modelo.variaveis) for modelo in modelos],
            "Equacao Ajustada": [modelo.equacao for modelo in modelos],
            "R2": [modelo.r2 for modelo in modelos],
            "R2 Ajustado": [modelo.r2_ajustado for modelo in modelos],
            "RMSE": [modelo.rmse for modelo in modelos],
            "F-stat": [modelo.f_estatistica for modelo in modelos],
            "p-valor(F)": [modelo.f_pvalor for modelo in modelos],
        }
    )


def escolher_melhor_modelo(modelos: list[ResultadoRegressao]) -> ResultadoRegressao:
    """Escolhe o melhor modelo por maior R2 ajustado e, em empate, menor RMSE."""
    return sorted(modelos, key=lambda modelo: (-modelo.r2_ajustado, modelo.rmse))[0]


def montar_tabelas_coeficientes(modelos: list[ResultadoRegressao]) -> dict[str, pd.DataFrame]:
    """Monta uma tabela de coeficientes para cada modelo ajustado."""
    tabelas = {}
    for modelo in modelos:
        tabelas[modelo.nome] = pd.DataFrame(
            {
                "Termo": modelo.nomes_coeficientes,
                "Estimativa": modelo.coeficientes,
                "Erro Padrao": modelo.erros_padrao,
                "t": modelo.t_estatisticas,
                "p-valor": modelo.p_valores,
            }
        )
    return tabelas


# ===========================================================
# PASSO 5 - ANALISE DE RESIDUOS DO MELHOR MODELO
# ===========================================================


def gerar_diagnosticos_residuos(df: pd.DataFrame, melhor_modelo: ResultadoRegressao) -> None:
    """Gera quatro graficos diagnosticos dos residuos em uma figura 2x2."""
    residuos = melhor_modelo.residuos
    ajustados = melhor_modelo.y_previsto

    fig, axes = plt.subplots(2, 2, figsize=(11, 8), dpi=150)

    axes[0, 0].scatter(ajustados, residuos, color="#2F75B5", alpha=0.75)
    axes[0, 0].axhline(0, color="#C00000", linestyle="--", linewidth=1.5)
    axes[0, 0].set_title("Resíduos vs Valores Ajustados", fontweight="bold")
    axes[0, 0].set_xlabel("Valores ajustados")
    axes[0, 0].set_ylabel("Resíduos")

    sns.histplot(residuos, kde=False, stat="density", bins=12, color="#BDD7EE", ax=axes[0, 1])
    media_res = np.mean(residuos)
    desvio_res = np.std(residuos, ddof=1)
    x_norm = np.linspace(residuos.min(), residuos.max(), 200)
    axes[0, 1].plot(x_norm, stats.norm.pdf(x_norm, media_res, desvio_res), color="#C00000")
    axes[0, 1].set_title("Histograma dos Resíduos", fontweight="bold")
    axes[0, 1].set_xlabel("Resíduos")

    stats.probplot(residuos, dist="norm", plot=axes[1, 0])
    axes[1, 0].set_title("QQ-Plot dos Resíduos", fontweight="bold")

    axes[1, 1].scatter(df["NOTAFG"], residuos, color="#2F75B5", alpha=0.75)
    axes[1, 1].axhline(0, color="#C00000", linestyle="--", linewidth=1.5)
    axes[1, 1].set_title("Resíduos vs NOTAFG", fontweight="bold")
    axes[1, 1].set_xlabel("NOTAFG")
    axes[1, 1].set_ylabel("Resíduos")

    for ax in axes.flat:
        ax.grid(True, linestyle="--", alpha=0.35)

    plt.tight_layout()
    plt.savefig(ARQUIVO_DIAGNOSTICO, dpi=300)
    plt.close()


def testar_normalidade_residuos(residuos: np.ndarray) -> tuple[float | None, float | None, str]:
    """Executa Shapiro-Wilk quando n < 5000 e retorna interpretacao textual."""
    n = len(residuos)
    if n >= 5000:
        interpretacao = (
            "Teste Shapiro-Wilk nao aplicado porque n >= 5000. "
            "Para amostras grandes, recomenda-se interpretar tambem QQ-Plot e histograma."
        )
        print(interpretacao)
        return None, None, interpretacao

    estatistica, p_valor = stats.shapiro(residuos)
    if p_valor >= 0.05:
        interpretacao = (
            "Como p-valor >= 0,05, nao ha evidencia estatistica suficiente "
            "para rejeitar a normalidade dos residuos."
        )
    else:
        interpretacao = (
            "Como p-valor < 0,05, ha evidencia estatistica de desvio "
            "da normalidade dos residuos."
        )

    print(f"Shapiro-Wilk: W = {estatistica:.4f}; p-valor = {p_valor:.4f}")
    print(interpretacao)
    return float(estatistica), float(p_valor), interpretacao


def criar_tabela_residuos(df: pd.DataFrame, melhor_modelo: ResultadoRegressao) -> pd.DataFrame:
    """Cria a tabela de residuos do melhor modelo."""
    return pd.DataFrame(
        {
            "Id": df["Id"],
            "NOTAFG": df["NOTAFG"],
            "DOUTORES": df["DOUTORES"],
            "NOTACE Real": df["NOTACE"],
            "NOTACE Previsto": melhor_modelo.y_previsto,
            "Residuo": melhor_modelo.residuos,
        }
    )


# ===========================================================
# PASSO 6 - PREVISAO PARA UM CURSO ALEATORIO
# ===========================================================


def vetor_x0_para_modelo(linha: pd.Series, modelo: ResultadoRegressao) -> np.ndarray:
    """Monta o vetor [1, x1, x2, ...] na ordem de coeficientes do modelo."""
    valores = [1.0]
    for variavel in modelo.variaveis:
        valores.append(float(linha[variavel]))
    return np.array(valores, dtype=float)


def criar_previsao_curso_sorteado(
    df: pd.DataFrame,
    melhor_modelo: ResultadoRegressao,
) -> tuple[pd.DataFrame, str]:
    """Sorteia um curso, calcula previsao, erro e intervalo de 95%."""
    curso = df.sample(1, random_state=42).iloc[0]
    x0 = vetor_x0_para_modelo(curso, melhor_modelo)

    valor_real = float(curso["NOTACE"])
    valor_previsto = float(x0 @ melhor_modelo.coeficientes)
    erro_absoluto = abs(valor_real - valor_previsto)
    erro_relativo = (erro_absoluto / valor_real) * 100 if valor_real != 0 else np.nan

    t_critico = stats.t.ppf(0.975, melhor_modelo.gl_residuais)
    erro_padrao_previsao = np.sqrt(melhor_modelo.mse * (1.0 + x0.T @ melhor_modelo.xtx_inv @ x0))
    limite_inferior = valor_previsto - t_critico * erro_padrao_previsao
    limite_superior = valor_previsto + t_critico * erro_padrao_previsao

    tabela = pd.DataFrame(
        {
            "Id": [curso["Id"]],
            "Curso": [curso.get("Curso", "Ciência da Computação")],
            "Codigo IES": [curso.get("Código IES", np.nan)],
            "NOTAFG": [curso["NOTAFG"]],
            "DOUTORES": [curso["DOUTORES"]],
            "NOTACE Real": [valor_real],
            "NOTACE Previsto": [valor_previsto],
            "Erro Absoluto": [erro_absoluto],
            "Erro Relativo (%)": [erro_relativo],
            "IC 95% Inferior": [limite_inferior],
            "IC 95% Superior": [limite_superior],
            "Melhor Modelo": [melhor_modelo.nome],
        }
    )

    interpretacao = (
        f"O curso sorteado apresentou NOTACE real de {valor_real:.4f}. "
        f"O modelo previu {valor_previsto:.4f}, com erro absoluto de "
        f"{erro_absoluto:.4f} e erro relativo de {erro_relativo:.2f}%. "
        f"O intervalo de previsão individual de 95% vai de "
        f"{limite_inferior:.4f} a {limite_superior:.4f}."
    )

    print("Previsao para curso sorteado:")
    print(tabela.to_string(index=False))
    print(interpretacao)
    return tabela, interpretacao


# ===========================================================
# PASSO 7 - EXPORTACAO E FORMATACAO DO EXCEL
# ===========================================================


def exportar_excel(
    matriz_correlacao: pd.DataFrame,
    modelos: list[ResultadoRegressao],
    comparacao: pd.DataFrame,
    melhor_modelo: ResultadoRegressao,
    tabelas_coeficientes: dict[str, pd.DataFrame],
    tabela_residuos: pd.DataFrame,
    shapiro_w: float | None,
    shapiro_p: float | None,
    interpretacao_shapiro: str,
    tabela_previsao: pd.DataFrame,
    interpretacao_previsao: str,
) -> None:
    """Exporta todas as abas pedidas e aplica formatacao com openpyxl."""
    with pd.ExcelWriter(ARQUIVO_EXCEL_SAIDA, engine="openpyxl") as writer:
        criar_aba_identificacao(writer)
        matriz_correlacao.to_excel(writer, sheet_name="1_Correlacao")
        comparacao.to_excel(writer, sheet_name="2_Modelos_Comparacao", index=False, startrow=0)

        linha_atual = len(comparacao) + 4
        for modelo in modelos:
            pd.DataFrame({"Tabela": [f"Coeficientes - {modelo.nome}"]}).to_excel(
                writer,
                sheet_name="2_Modelos_Comparacao",
                index=False,
                header=False,
                startrow=linha_atual,
            )
            tabelas_coeficientes[modelo.nome].to_excel(
                writer,
                sheet_name="2_Modelos_Comparacao",
                index=False,
                startrow=linha_atual + 1,
            )
            linha_atual += len(tabelas_coeficientes[modelo.nome]) + 4

        conclusao = (
            f"Conclusão: o melhor modelo é {melhor_modelo.nome}, pois apresenta "
            f"o maior R2 ajustado ({melhor_modelo.r2_ajustado:.4f}) e RMSE "
            f"igual a {melhor_modelo.rmse:.4f}. Em regressão, o R2 ajustado "
            "favorece modelos com bom poder explicativo sem premiar "
            "indevidamente a inclusão de variáveis."
        )
        pd.DataFrame({"Conclusao": [conclusao]}).to_excel(
            writer,
            sheet_name="2_Modelos_Comparacao",
            index=False,
            startrow=linha_atual,
        )

        tabela_residuos.to_excel(writer, sheet_name="3_Residuos", index=False, startrow=0)
        linha_resumo_residuos = len(tabela_residuos) + 3
        pd.DataFrame(
            {
                "Teste": ["Shapiro-Wilk"],
                "Estatistica W": [shapiro_w],
                "p-valor": [shapiro_p],
                "Interpretacao": [interpretacao_shapiro],
            }
        ).to_excel(
            writer,
            sheet_name="3_Residuos",
            index=False,
            startrow=linha_resumo_residuos,
        )

        tabela_previsao.to_excel(writer, sheet_name="4_Previsao", index=False, startrow=0)
        pd.DataFrame({"Interpretacao": [interpretacao_previsao]}).to_excel(
            writer,
            sheet_name="4_Previsao",
            index=False,
            startrow=4,
        )

    formatar_excel(
        comparacao=comparacao,
        tabela_residuos=tabela_residuos,
        linha_previsao=2,
    )


def criar_aba_identificacao(writer: pd.ExcelWriter) -> None:
    """Cria a primeira aba com identificacao obrigatoria do projeto."""
    dados_identificacao = pd.DataFrame(
        {
            "Campo": [
                "Projeto CPC 2021 — Análise de Regressão",
                "Curso Analisado",
                "Disciplina",
                "Professor",
                "Data de Entrega",
                "Integrantes da Equipe",
                "[NOME 1]",
                "[NOME 2]",
                "[NOME 3]",
                "[NOME 4]",
            ],
            "Informacao": [
                "",
                "Ciência da Computação (Bacharelado)",
                "[nome da disciplina]",
                "[nome do professor]",
                DATA_ENTREGA,
                "",
                "",
                "",
                "",
                "",
            ],
        }
    )
    dados_identificacao.to_excel(
        writer,
        sheet_name="0_Identificacao",
        index=False,
        header=False,
    )


def formatar_excel(comparacao: pd.DataFrame, tabela_residuos: pd.DataFrame, linha_previsao: int) -> None:
    """Aplica formatacao padrao, destaques e imagens no arquivo Excel."""
    workbook = load_workbook(ARQUIVO_EXCEL_SAIDA)

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        aplicar_fonte_padrao(worksheet)
        aplicar_bordas_e_linhas_alternadas(worksheet)
        ajustar_larguras(worksheet)

    formatar_aba_identificacao(workbook["0_Identificacao"])

    for nome_aba in ["1_Correlacao", "2_Modelos_Comparacao", "3_Residuos", "4_Previsao"]:
        worksheet = workbook[nome_aba]
        worksheet.freeze_panes = "A2"
        formatar_cabecalho(worksheet, 1)
        formatar_numeros(worksheet)

    destacar_melhores_metricas(workbook["2_Modelos_Comparacao"], comparacao)
    workbook["4_Previsao"][linha_previsao][0].fill = PatternFill("solid", fgColor=AMARELO)
    for cell in workbook["4_Previsao"][linha_previsao]:
        cell.fill = PatternFill("solid", fgColor=AMARELO)

    inserir_imagens_excel(workbook)

    workbook.save(ARQUIVO_EXCEL_SAIDA)


def aplicar_fonte_padrao(worksheet) -> None:
    """Define fonte Arial nas celulas usadas."""
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def aplicar_bordas_e_linhas_alternadas(worksheet) -> None:
    """Aplica bordas finas e alternancia branco/azul claro."""
    borda = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    preenchimento_claro = PatternFill("solid", fgColor=AZUL_CLARO)
    preenchimento_branco = PatternFill("solid", fgColor=BRANCO)

    for row in worksheet.iter_rows():
        if all(cell.value is None for cell in row):
            continue
        for cell in row:
            cell.border = borda
            if cell.row > 1:
                cell.fill = preenchimento_claro if cell.row % 2 == 0 else preenchimento_branco


def formatar_cabecalho(worksheet, linha: int) -> None:
    """Formata uma linha como cabecalho azul escuro."""
    preenchimento = PatternFill("solid", fgColor=AZUL_ESCURO)
    fonte = Font(name="Arial", size=10, color=BRANCO, bold=True)
    for cell in worksheet[linha]:
        if cell.value is not None:
            cell.fill = preenchimento
            cell.font = fonte
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def formatar_aba_identificacao(worksheet) -> None:
    """Aplica formatacao especial da primeira aba obrigatoria."""
    worksheet.freeze_panes = "A1"
    worksheet.merge_cells("A1:B1")
    worksheet["A1"].font = Font(name="Arial", size=12, color=BRANCO, bold=True)
    worksheet["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, 3):
            cell = worksheet.cell(row=row, column=col)
            cell.font = Font(name="Arial", size=12, bold=(row == 6))
            cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 48


def formatar_numeros(worksheet) -> None:
    """Formata valores numericos com quatro casas decimais."""
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def ajustar_larguras(worksheet) -> None:
    """Ajusta largura das colunas conforme conteudo."""
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 55)


def destacar_melhores_metricas(worksheet, comparacao: pd.DataFrame) -> None:
    """Destaca em verde o melhor modelo de cada metrica numerica."""
    preenchimento_verde = PatternFill("solid", fgColor=VERDE)

    metricas_maior_melhor = ["R2", "R2 Ajustado", "F-stat"]
    metricas_menor_melhor = ["RMSE", "p-valor(F)"]

    for metrica in metricas_maior_melhor:
        col_idx = comparacao.columns.get_loc(metrica) + 1
        valor_melhor = comparacao[metrica].max()
        for row_idx, valor in enumerate(comparacao[metrica], start=2):
            if np.isclose(valor, valor_melhor):
                worksheet.cell(row=row_idx, column=col_idx).fill = preenchimento_verde

    for metrica in metricas_menor_melhor:
        col_idx = comparacao.columns.get_loc(metrica) + 1
        valor_melhor = comparacao[metrica].min()
        for row_idx, valor in enumerate(comparacao[metrica], start=2):
            if np.isclose(valor, valor_melhor):
                worksheet.cell(row=row_idx, column=col_idx).fill = preenchimento_verde


def inserir_imagens_excel(workbook) -> None:
    """Insere os graficos salvos nas abas correspondentes do Excel."""
    imagens = [
        ("1_Correlacao", ARQUIVO_HEATMAP, "E2", 420, 300),
        ("1_Correlacao", ARQUIVO_SCATTER_FG, "E20", 420, 300),
        ("1_Correlacao", ARQUIVO_SCATTER_DOUTORES, "M20", 420, 300),
        ("3_Residuos", ARQUIVO_DIAGNOSTICO, "H2", 650, 470),
    ]

    for nome_aba, caminho, ancora, largura, altura in imagens:
        if not caminho.exists():
            continue
        imagem = ExcelImage(str(caminho))
        imagem.width = largura
        imagem.height = altura
        workbook[nome_aba].add_image(imagem, ancora)


# ===========================================================
# EXECUCAO PRINCIPAL
# ===========================================================


def main() -> None:
    """Executa toda a analise de regressao e gera o Excel final."""
    df = carregar_e_limpar_dados()

    matriz_correlacao = gerar_graficos_correlacao(df)

    modelos = ajustar_todos_os_modelos(df)
    comparacao = criar_tabela_comparacao(modelos)
    melhor_modelo = escolher_melhor_modelo(modelos)
    tabelas_coeficientes = montar_tabelas_coeficientes(modelos)

    print("\nTabela comparativa dos modelos:")
    print(comparacao.to_string(index=False))
    print(f"\nMelhor modelo selecionado: {melhor_modelo.nome}")

    gerar_diagnosticos_residuos(df, melhor_modelo)
    shapiro_w, shapiro_p, interpretacao_shapiro = testar_normalidade_residuos(
        melhor_modelo.residuos
    )
    tabela_residuos = criar_tabela_residuos(df, melhor_modelo)

    tabela_previsao, interpretacao_previsao = criar_previsao_curso_sorteado(df, melhor_modelo)

    exportar_excel(
        matriz_correlacao=matriz_correlacao,
        modelos=modelos,
        comparacao=comparacao,
        melhor_modelo=melhor_modelo,
        tabelas_coeficientes=tabelas_coeficientes,
        tabela_residuos=tabela_residuos,
        shapiro_w=shapiro_w,
        shapiro_p=shapiro_p,
        interpretacao_shapiro=interpretacao_shapiro,
        tabela_previsao=tabela_previsao,
        interpretacao_previsao=interpretacao_previsao,
    )

    print(f"\nArquivo Excel gerado: {ARQUIVO_EXCEL_SAIDA.resolve()}")
    print("Graficos gerados:")
    print(f"- {ARQUIVO_HEATMAP.resolve()}")
    print(f"- {ARQUIVO_SCATTER_FG.resolve()}")
    print(f"- {ARQUIVO_SCATTER_DOUTORES.resolve()}")
    print(f"- {ARQUIVO_DIAGNOSTICO.resolve()}")


if __name__ == "__main__":
    main()
